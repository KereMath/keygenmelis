import openpyxl
import os

def delete_empty_sheets(file_path):
    try:
        # Open the workbook
        wb = openpyxl.load_workbook(file_path)

        # List to store sheets to be removed
        sheets_to_remove = []

        # Iterate over sheets
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            if ws.max_row == 1 and ws.max_column == 1 and ws.cell(1, 1).value is None:
                sheets_to_remove.append(sheet)

        # Remove empty sheets
        for sheet in sheets_to_remove:
            wb.remove(wb[sheet])

        # Save the workbook
        wb.save(file_path)
        print("Empty sheets removed successfully.")
    except Exception as e:
        print(f"An error occurred: {e}")

if __name__ == "__main__":
    file_path = os.path.expanduser("~/Desktop/testResults.xlsx")
    if os.path.exists(file_path):
        delete_empty_sheets(file_path)
    else:
        print(f"File not found: {file_path}")
