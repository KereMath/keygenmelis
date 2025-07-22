import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

def process_excel_files(folder_path, output_file):
    # 0.01 ve 0.05 tablolarını tutacak DataFrame'ler
    table_001 = []
    table_005 = []

    # Klasördeki tüm Excel dosyalarını bul
    files = [file for file in os.listdir(folder_path) if file.endswith('.xlsx')]

    if not files:
        print("Klasörde Excel dosyası bulunamadı.")
        return

    for file in files:
        file_path = os.path.join(folder_path, file)

        # Dosya adından Linear Complexity ve tür bilgilerini çıkar
        if "_LC0.01" in file:
            complexity = 0.01
        elif "_LC0.05" in file:
            complexity = 0.05
        else:
            continue

        # Tür bilgisini çek
        if "aes" in file:
            data_type = "aes"
        elif "shrinking" in file:
            data_type = "shrinking"
        elif "alternating" in file:
            data_type = "alternating"
        else:
            data_type = "unknown"

        # Bit boyutunu dosya adından al
        bit_size = int(file.split('-')[-1].split('_')[0])

        # Excel dosyasını oku
        df = pd.read_excel(file_path)

        # Her test için Passed ve Failed sayısını hesapla
        test_results = {"Bit Size": bit_size, "Type": data_type}
        for column in df.columns[1:]:
            passed_count = df[column].str.count("Passed").sum()
            failed_count = df[column].str.count("Failed").sum()
            test_results[column] = (passed_count, failed_count)

        # Sonuçları ilgili tabloya ekle
        if complexity == 0.01:
            table_001.append(test_results)
        elif complexity == 0.05:
            table_005.append(test_results)

    # DataFrame'leri oluştur
    table_001 = pd.DataFrame(table_001)
    table_005 = pd.DataFrame(table_005)

    # Excel dosyasını yaz
    write_grouped_excel(output_file, "0.01", table_001)
    write_grouped_excel(output_file, "0.05", table_005, mode="a")


def write_grouped_excel(output_file, sheet_name, data, mode="w"):
    # Excel çalışma kitabı oluştur veya var olanı yükle
    if mode == "w":
        wb = Workbook()
    else:
        wb = load_workbook(output_file)

    ws = wb.create_sheet(title=sheet_name)

    # Başlıkları ayarla
    ws.append(["Bit Size", "Type", "Frequency Test", "Run Count Test", "Run L1 Test",
               "Template Match 4-1 Test", "Template Match 4-2 Test",
               "Template Match 4-3 Test", "Template Match 4-4 Test",
               "Linear Complexity Test", "Blind Spot Complexity Test"])

    # Her Bit Size için verileri ekle
    grouped = data.groupby("Bit Size")
    for bit_size, group in grouped:
        # Türlere göre alt satırları gruplandır
        aes = group[group["Type"] == "aes"]
        shrinking = group[group["Type"] == "shrinking"]
        alternating = group[group["Type"] == "alternating"]

        # Bit Size için ilk satır
        ws.append([bit_size, ""])

        # AES, Shrinking ve Alternating sonuçlarını ekle
        for row_data, label in zip([aes, shrinking, alternating], ["AES", "Shrinking", "Alternating"]):
            row = ["", label]  # Bit Size yalnızca ilk satırda gösterilir
            for test in data.columns[2:]:
                test_result = row_data[test].values[0] if len(row_data) > 0 else (0, 0)
                row.append(f"{test_result[0]}/{test_result[1]}")
            ws.append(row)

    # Stiller ve hizalamalar
    for col in ws.columns:
        max_length = max(len(str(cell.value) if cell.value else "") for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = max_length + 2
        for cell in col:
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if cell.row == 1:
                cell.font = Font(bold=True)

    # Çalışma kitabını kaydet
    wb.save(output_file)


if __name__ == "__main__":
    folder_path = r"C:\Users\J.A.R.V.I.S\Desktop\testExcel"
    output_file = r"C:\Users\J.A.R.V.I.S\Desktop\testResults.xlsx"
    process_excel_files(folder_path, output_file)
